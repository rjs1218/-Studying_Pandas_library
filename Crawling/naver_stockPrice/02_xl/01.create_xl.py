import openpyxl

# 엑셀 만들기
wb = openpyxl.Workbook()

# 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 엑셀 저장하기 / 문자열 앞에 r을 붙이면 문자열을 모두 문자로 취급해라 라는 뜻을 가짐
wb.save(r'/Users/gungo/Documents/Python/Crawling/naver_stockPrice/02.xl/참가자_data.xlsx')